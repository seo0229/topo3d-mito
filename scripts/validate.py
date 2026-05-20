import os
import numpy as np
import torch
from skimage.morphology import skeletonize
from scipy import ndimage
import tifffile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ------------------ Original Evaluation Metrics ------------------ #

def compute_dice(pred, gt):
    pred = pred.astype(bool)
    gt = gt.astype(bool)
    intersection = np.logical_and(pred, gt).sum()
    size_pred = pred.sum()
    size_gt = gt.sum()
    if size_pred + size_gt == 0:
        return 1.0
    return (2.0 * intersection) / (size_pred + size_gt)

def soft_dice(y_true, y_pred):
    smooth = 1.
    intersection = torch.sum(y_true * y_pred)
    return 1 - ((2. * intersection + smooth) /
                (torch.sum(y_true) + torch.sum(y_pred) + smooth))

def cl_score(v, s):
    return np.sum(np.logical_and(v, s)) / (np.sum(s) + 1e-8)

def clDice(v_p, v_l):
    s_l = skeletonize(v_l)
    s_p = skeletonize(v_p)
    tprec = cl_score(v_p, s_l)
    tsens = cl_score(v_l, s_p)
    return 2 * tprec * tsens / (tprec + tsens + 1e-8)

def compute_betti_0(pred, gt):
    _, num_pred = ndimage.label(pred)
    _, num_gt = ndimage.label(gt)
    return abs(num_pred - num_gt)

def IoU(pred, gt):
    intersection = np.logical_and(pred, gt).sum()
    union = np.logical_or(pred, gt).sum()
    if union == 0:
        return 0.0
    return intersection / union

# ------------------ Connectivity Test Function ------------------ #

def test_connectivity(pred_binary, gt_binary, num_samples=100, seed=None, max_distance=100):
    if seed is not None:
        np.random.seed(seed)
    
    pred_labels, num_pred_components = ndimage.label(pred_binary)
    gt_labels, num_gt_components = ndimage.label(gt_binary)
    
    gt_coords = np.argwhere(gt_binary > 0)
    pred_coords = np.argwhere(pred_binary > 0)
    
    if len(gt_coords) < 2 or len(pred_coords) < 2:
        return {
            'num_samples': 0,
            'num_passed': 0,
            'num_failed': 0,
            'accuracy': 0.0,
            'true_positives': 0,
            'true_negatives': 0,
            'false_positives': 0,
            'false_negatives': 0,
            'avg_distance': 0.0,
            'num_retries': 0,
            'sample_records': [],
            'error': 'Insufficient foreground points'
        }
    
    passed = 0
    failed = 0
    true_positives = 0
    true_negatives = 0
    false_positives = 0
    false_negatives = 0
    total_retries = 0
    distances = []
    samples_tested = 0
    sample_records = []  # NEW: store per-sample data for Excel export
    
    for _ in range(num_samples):
        max_retries = 100
        valid_pair_found = False
        
        for retry in range(max_retries):
            idx1, idx2 = np.random.choice(len(gt_coords), size=2, replace=False)
            point1 = tuple(gt_coords[idx1])
            point2 = tuple(gt_coords[idx2])
            distance = np.linalg.norm(gt_coords[idx1] - gt_coords[idx2])
            
            if distance <= max_distance:
                valid_pair_found = True
                distances.append(distance)
                total_retries += retry
                break
        
        if not valid_pair_found:
            continue
        
        samples_tested += 1
        
        point1_in_pred = pred_binary[point1] > 0
        point2_in_pred = pred_binary[point2] > 0
        
        gt_label1 = gt_labels[point1]
        gt_label2 = gt_labels[point2]
        connected_in_gt = (gt_label1 == gt_label2) and (gt_label1 > 0)
        
        if point1_in_pred and point2_in_pred:
            pred_label1 = pred_labels[point1]
            pred_label2 = pred_labels[point2]
            connected_in_pred = (pred_label1 == pred_label2) and (pred_label1 > 0)
        else:
            connected_in_pred = False
        
        if connected_in_gt and connected_in_pred:
            passed += 1
            true_positives += 1
            outcome = 'TP'
        elif not connected_in_gt and not connected_in_pred:
            passed += 1
            true_negatives += 1
            outcome = 'TN'
        elif not connected_in_gt and connected_in_pred:
            failed += 1
            false_positives += 1
            outcome = 'FP'
        elif connected_in_gt and not connected_in_pred:
            failed += 1
            false_negatives += 1
            outcome = 'FN'

        # NEW: record full details for this sample
        sample_records.append({
            'sample_index': samples_tested,
            'point1': point1,
            'point2': point2,
            'point1_in_pred': point1_in_pred,
            'point2_in_pred': point2_in_pred,
            'connected_in_gt': connected_in_gt,
            'connected_in_pred': connected_in_pred,
            'distance': distance,
            'outcome': outcome,
            'passed': outcome in ('TP', 'TN'),
        })
    
    accuracy = passed / samples_tested if samples_tested > 0 else 0.0
    avg_distance = np.mean(distances) if distances else 0.0
    
    return {
        'num_samples': samples_tested,
        'num_passed': passed,
        'num_failed': failed,
        'accuracy': accuracy,
        'true_positives': true_positives,
        'true_negatives': true_negatives,
        'false_positives': false_positives,
        'false_negatives': false_negatives,
        'num_pred_components': num_pred_components,
        'num_gt_components': num_gt_components,
        'avg_distance': avg_distance,
        'num_retries': total_retries,
        'sample_records': sample_records,  # NEW
    }

# ------------------ Excel Export Function ------------------ #

def export_to_excel(connectivity_results, pred_files, original_results, output_path):
    wb = Workbook()

    # ---- Styles ----
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill('solid', start_color='2F5496')
    sub_header_font = Font(name='Arial', bold=True, size=10)
    sub_header_fill = PatternFill('solid', start_color='D9E1F2')
    normal_font = Font(name='Arial', size=10)
    center = Alignment(horizontal='center', vertical='center')
    left = Alignment(horizontal='left', vertical='center')

    tp_fill = PatternFill('solid', start_color='C6EFCE')   # green
    tn_fill = PatternFill('solid', start_color='DDEBF7')   # blue
    fp_fill = PatternFill('solid', start_color='FFEB9C')   # yellow
    fn_fill = PatternFill('solid', start_color='FFC7CE')   # red
    pass_font = Font(name='Arial', size=10, color='375623')
    fail_font = Font(name='Arial', size=10, color='9C0006')

    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    outcome_fills = {'TP': tp_fill, 'TN': tn_fill, 'FP': fp_fill, 'FN': fn_fill}
    outcome_desc = {
        'TP': 'Connected in both GT and Pred',
        'TN': 'Disconnected in both GT and Pred',
        'FP': 'Wrongly connected in Pred',
        'FN': 'Wrongly disconnected in Pred',
    }

    def style_cell(cell, font=None, fill=None, alignment=None, border_=True):
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border_:
            cell.border = border

    def write_header_row(ws, row, cols):
        for col_idx, label in enumerate(cols, start=1):
            c = ws.cell(row=row, column=col_idx, value=label)
            style_cell(c, font=header_font, fill=header_fill, alignment=center)

    # ================================================================
    # Sheet 1: Summary
    # ================================================================
    ws_summary = wb.active
    ws_summary.title = 'Summary'

    ws_summary.merge_cells('A1:F1')
    title_cell = ws_summary['A1']
    title_cell.value = 'Connectivity Evaluation — Summary'
    title_cell.font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    title_cell.fill = PatternFill('solid', start_color='1F3864')
    title_cell.alignment = center
    ws_summary.row_dimensions[1].height = 24

    # Metrics header
    metrics_headers = ['Volume', 'Dice', 'clDice', 'Betti-0 Diff', 'IoU',
                        'Conn. Accuracy', 'TP', 'TN', 'FP', 'FN',
                        'GT Components', 'Pred Components', 'Avg Distance']
    write_header_row(ws_summary, 3, metrics_headers)

    for i, ((dice, cldice_score, betti, iou), conn) in enumerate(zip(original_results, connectivity_results)):
        r = 4 + i
        row_data = [
            os.path.basename(pred_files[i]),
            round(dice, 4), round(cldice_score, 4), betti, round(iou, 4),
            round(conn['accuracy'], 4),
            conn['true_positives'], conn['true_negatives'],
            conn['false_positives'], conn['false_negatives'],
            conn.get('num_gt_components', ''),
            conn.get('num_pred_components', ''),
            round(conn['avg_distance'], 2),
        ]
        for col_idx, val in enumerate(row_data, start=1):
            c = ws_summary.cell(row=r, column=col_idx, value=val)
            c.font = normal_font
            c.border = border
            c.alignment = center if col_idx > 1 else left

    # Averages row
    n = len(original_results)
    avg_row = 4 + n + 1
    ws_summary.cell(row=avg_row, column=1, value='AVERAGE').font = sub_header_font
    ws_summary['A' + str(avg_row)].fill = sub_header_fill

    avg_cols = [2, 3, 4, 5, 6]  # Dice, clDice, Betti, IoU, Conn Acc
    for col in avg_cols:
        col_letter = get_column_letter(col)
        formula = f'=AVERAGE({col_letter}4:{col_letter}{4+n-1})'
        c = ws_summary.cell(row=avg_row, column=col, value=formula)
        c.font = sub_header_font
        c.border = border
        c.alignment = center
        c.number_format = '0.0000'

    # Column widths
    col_widths = [32, 10, 10, 13, 10, 16, 7, 7, 7, 7, 16, 18, 16]
    for i, w in enumerate(col_widths, start=1):
        ws_summary.column_dimensions[get_column_letter(i)].width = w

    # Legend
    legend_row = avg_row + 3
    ws_summary.cell(row=legend_row, column=1, value='Legend').font = sub_header_font
    legend_items = [('TP — True Positive', tp_fill), ('TN — True Negative', tn_fill),
                    ('FP — False Positive', fp_fill), ('FN — False Negative', fn_fill)]
    for j, (label, fill) in enumerate(legend_items):
        c = ws_summary.cell(row=legend_row + 1 + j, column=1, value=label)
        c.font = normal_font
        c.fill = fill
        c.border = border

    # ================================================================
    # Sheet per volume: sampled coordinates
    # ================================================================
    coord_headers = [
        'Sample #',
        'Point1 Z', 'Point1 Y', 'Point1 X',
        'Point2 Z', 'Point2 Y', 'Point2 X',
        'Distance',
        'Connected in GT?',
        'Point1 in Pred?', 'Point2 in Pred?',
        'Connected in Pred?',
        'Outcome', 'Outcome Description', 'Passed?',
    ]

    for i, (conn, pred_path) in enumerate(zip(connectivity_results, pred_files)):
        sheet_name = f'Vol{i+1}_Coords'
        ws = wb.create_sheet(title=sheet_name)

        # Title
        ws.merge_cells(f'A1:{get_column_letter(len(coord_headers))}1')
        t = ws['A1']
        t.value = f'Sampled Point Pairs — {os.path.basename(pred_path)}'
        t.font = Font(name='Arial', bold=True, size=12, color='FFFFFF')
        t.fill = PatternFill('solid', start_color='1F3864')
        t.alignment = center
        ws.row_dimensions[1].height = 22

        # Sub-info
        ws.cell(row=2, column=1,
                value=f"Connectivity Accuracy: {conn['accuracy']*100:.2f}%  |  "
                      f"GT Components: {conn.get('num_gt_components','')}  |  "
                      f"Pred Components: {conn.get('num_pred_components','')}  |  "
                      f"Avg Distance: {conn['avg_distance']:.2f} voxels")
        ws.cell(row=2, column=1).font = Font(name='Arial', italic=True, size=9, color='595959')

        write_header_row(ws, 3, coord_headers)

        for rec in conn.get('sample_records', []):
            p1 = rec['point1']
            p2 = rec['point2']
            row_vals = [
                rec['sample_index'],
                p1[0] if len(p1) > 2 else '',   # Z (may not exist for 2D)
                p1[-2] if len(p1) >= 2 else p1[0],
                p1[-1],
                p2[0] if len(p2) > 2 else '',
                p2[-2] if len(p2) >= 2 else p2[0],
                p2[-1],
                round(rec['distance'], 2),
                'Yes' if rec['connected_in_gt'] else 'No',
                'Yes' if rec['point1_in_pred'] else 'No',
                'Yes' if rec['point2_in_pred'] else 'No',
                'Yes' if rec['connected_in_pred'] else 'No',
                rec['outcome'],
                outcome_desc[rec['outcome']],
                'PASS' if rec['passed'] else 'FAIL',
            ]
            r = 3 + rec['sample_index']
            fill = outcome_fills[rec['outcome']]
            for col_idx, val in enumerate(row_vals, start=1):
                c = ws.cell(row=r, column=col_idx, value=val)
                c.font = normal_font
                c.fill = fill
                c.border = border
                c.alignment = center

            # Color the Passed? cell distinctly
            pass_cell = ws.cell(row=r, column=len(coord_headers))
            pass_cell.font = pass_font if rec['passed'] else fail_font

        # Column widths
        widths = [10, 10, 10, 10, 10, 10, 10, 12, 18, 16, 16, 20, 10, 32, 10]
        for ci, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.freeze_panes = 'A4'

    wb.save(output_path)
    print(f"\nExcel file saved to: {output_path}")

# ------------------ File Lists ------------------ #
pred_files = [
            '/data/elliott/baselines/pytorch_connectomics/outputs/baseline/3D-GTO/test/prediction_000.tiff',
            '/data/elliott/baselines/pytorch_connectomics/outputs/baseline/3D-GTO/test/prediction_001.tiff',
            '/data/elliott/baselines/pytorch_connectomics/outputs/baseline/3D-GTO/test/prediction_002.tiff',
            '/data/elliott/baselines/pytorch_connectomics/outputs/baseline/3D-GTO/test/prediction_003.tiff',
            '/data/elliott/baselines/pytorch_connectomics/outputs/baseline/3D-GTO/test/prediction_004.tiff',
            '/data/elliott/baselines/pytorch_connectomics/outputs/baseline/3D-GTO/test/prediction_005.tiff',
]

gt_files = [
    '/data/elliott/baselines/pytorch_connectomics/datasets/MitoAICS/allfiles/3Dedited/raw1.tif',
    '/data/elliott/baselines/pytorch_connectomics/datasets/MitoAICS/allfiles/3Dedited/raw4.tif',
    '/data/elliott/baselines/pytorch_connectomics/datasets/MitoAICS/allfiles/3Dedited/raw5.tif',
    '/data/elliott/baselines/pytorch_connectomics/datasets/MitoAICS/allfiles/3Dedited/raw7.tif',
    '/data/elliott/baselines/pytorch_connectomics/datasets/MitoAICS/allfiles/3Dedited/raw28.tif',
    '/data/elliott/baselines/pytorch_connectomics/datasets/MitoAICS/allfiles/3Dedited/raw29.tif',
]

# ------------------ Output Paths ------------------ 
output_dir = "/data/elliott/baselines/pytorch_connectomics/outputs/baseline/3D-GTO/test/results"
os.makedirs(output_dir, exist_ok=True)
connectivity_file = os.path.join(output_dir, "connectivity_metrics.txt")
combined_metrics_file = os.path.join(output_dir, "combined_metrics.txt")
excel_file = os.path.join(output_dir, "connectivity_coordinates.xlsx")  # NEW
masks_dir = os.path.join(output_dir, "pred_masks")
os.makedirs(masks_dir, exist_ok=True)

# ------------------ Connectivity Testing Loop ------------------ #
print("="*60)
print("CONNECTIVITY TESTING + ORIGINAL METRICS")
print("="*60)

connectivity_results = []
all_accuracies = []
dice_scores, cldice_scores, betti_scores, iou_scores = [], [], [], []
original_results = []

for i, (pred_path, gt_path) in enumerate(zip(pred_files, gt_files)):
    print(f"\n--- Testing pair {i+1} ---")
    print(f"Prediction: {os.path.basename(pred_path)}")
    print(f"Ground Truth: {os.path.basename(gt_path)}")

    pred = tifffile.imread(pred_path)
    gt = tifffile.imread(gt_path)

    pred = np.squeeze(pred).astype(np.float32) / 255
    gt = np.squeeze(gt).astype(np.float32)
    
    pred_binary = (pred > 0.5).astype(np.uint8)
    gt_binary = (gt > 0.5).astype(np.uint8)

    mask_filename = os.path.splitext(os.path.basename(pred_path))[0] + "_mask.tiff"
    tifffile.imwrite(os.path.join(masks_dir, mask_filename), pred_binary)

    print(f"\nComputing original metrics...")
    dice = compute_dice(pred_binary, gt_binary)
    cldice_score = clDice(pred_binary, gt_binary)
    betti = compute_betti_0(pred_binary, gt_binary)
    iou = IoU(pred_binary, gt_binary)

    print(f'Dice Coefficient: {dice:.4f}')
    print(f'clDice: {cldice_score:.4f}')
    print(f'Betti-0 Difference: {betti}')
    print(f'IoU: {iou:.4f}')

    original_results.append((dice, cldice_score, betti, iou))
    dice_scores.append(dice)
    cldice_scores.append(cldice_score)
    betti_scores.append(betti)
    iou_scores.append(iou)

    print(f"\nTesting 100 random point pairs for connectivity...")
    conn_stats = test_connectivity(pred_binary, gt_binary, num_samples=100, seed=10+i)
    
    connectivity_results.append(conn_stats)
    if 'error' not in conn_stats:
        all_accuracies.append(conn_stats['accuracy'])
    
    print(f"\nConnectivity Test Results:")
    print(f"  Samples tested: {conn_stats['num_samples']}")
    print(f"  Passed: {conn_stats['num_passed']}")
    print(f"  Failed: {conn_stats['num_failed']}")
    print(f"  Accuracy: {conn_stats['accuracy']:.4f} ({conn_stats['accuracy']*100:.2f}%)")
    print(f"\nDetailed Breakdown:")
    print(f"  True Positives (connected in both): {conn_stats['true_positives']}")
    print(f"  True Negatives (disconnected in both): {conn_stats['true_negatives']}")
    print(f"  False Positives (wrongly connected): {conn_stats['false_positives']}")
    print(f"  False Negatives (wrongly disconnected): {conn_stats['false_negatives']}")
    print(f"\nComponent Counts:")
    print(f"  GT components: {conn_stats['num_gt_components']}")
    print(f"  Pred components: {conn_stats['num_pred_components']}")

# ------------------ Compute Averages ------------------ #
avg_dice = np.mean(dice_scores)
avg_cldice = np.mean(cldice_scores)
avg_betti = np.mean(betti_scores)
avg_iou = np.mean(iou_scores)

std_dice = np.std(dice_scores)
std_cldice = np.std(cldice_scores)
std_betti = np.std(betti_scores)
std_iou = np.std(iou_scores)

# ------------------ Save Results to Files ------------------ #

with open(connectivity_file, "w") as f:
    f.write("="*60 + "\n")
    f.write("CONNECTIVITY TEST RESULTS\n")
    f.write("="*60 + "\n")
    f.write(f"Testing method: Sample 100 random point pairs per volume\n")
    f.write(f"Connectivity criterion: Points are connected if in same 3D component\n\n")
    
    for i, (conn_stats, pred_path) in enumerate(zip(connectivity_results, pred_files)):
        f.write(f"\n--- Pair {i+1}: {os.path.basename(pred_path)} ---\n")
        f.write(f"Samples tested: {conn_stats['num_samples']}\n")
        f.write(f"Passed: {conn_stats['num_passed']}\n")
        f.write(f"Failed: {conn_stats['num_failed']}\n")
        f.write(f"Accuracy: {conn_stats['accuracy']:.4f} ({conn_stats['accuracy']*100:.2f}%)\n")
        f.write(f"\nDetailed Breakdown:\n")
        f.write(f"  True Positives (connected in both): {conn_stats['true_positives']}\n")
        f.write(f"  True Negatives (disconnected in both): {conn_stats['true_negatives']}\n")
        f.write(f"  False Positives (wrongly connected): {conn_stats['false_positives']}\n")
        f.write(f"  False Negatives (wrongly disconnected): {conn_stats['false_negatives']}\n")
        f.write(f"\nComponent Counts:\n")
        f.write(f"  GT components: {conn_stats['num_gt_components']}\n")
        f.write(f"  Pred components: {conn_stats['num_pred_components']}\n")
        f.write("-"*60 + "\n")
    
    if all_accuracies:
        f.write(f"\n{'='*60}\n")
        f.write(f"OVERALL CONNECTIVITY STATISTICS\n")
        f.write(f"{'='*60}\n")
        f.write(f"Average Connectivity Accuracy: {np.mean(all_accuracies):.4f} ± {np.std(all_accuracies):.4f}\n")
        f.write(f"Min Accuracy: {np.min(all_accuracies):.4f}\n")
        f.write(f"Max Accuracy: {np.max(all_accuracies):.4f}\n")

with open(combined_metrics_file, "w") as f:
    f.write("="*60 + "\n")
    f.write("COMBINED EVALUATION RESULTS\n")
    f.write("="*60 + "\n\n")
    
    for i, (dice, cldice_score, betti, iou) in enumerate(original_results):
        conn_stats = connectivity_results[i]
        f.write(f"\n--- Pair {i+1}: {os.path.basename(pred_files[i])} ---\n")
        f.write(f"\nOriginal Metrics:\n")
        f.write(f"  Dice Coefficient: {dice:.4f}\n")
        f.write(f"  clDice: {cldice_score:.4f}\n")
        f.write(f"  Betti-0 Difference: {betti}\n")
        f.write(f"  IoU: {iou:.4f}\n")
        f.write(f"\nConnectivity Metrics:\n")
        f.write(f"  Connectivity Accuracy: {conn_stats['accuracy']:.4f} ({conn_stats['accuracy']*100:.2f}%)\n")
        f.write(f"  Passed/Failed: {conn_stats['num_passed']}/{conn_stats['num_failed']} (out of {conn_stats['num_samples']})\n")
        f.write(f"  True Positives: {conn_stats['true_positives']}\n")
        f.write(f"  True Negatives: {conn_stats['true_negatives']}\n")
        f.write(f"  False Positives: {conn_stats['false_positives']}\n")
        f.write(f"  False Negatives: {conn_stats['false_negatives']}\n")
        f.write("-"*60 + "\n")
    
    f.write(f"\n{'='*60}\n")
    f.write(f"OVERALL AVERAGES\n")
    f.write(f"{'='*60}\n")
    f.write(f"\nOriginal Metrics:\n")
    f.write(f"  Average Dice Coefficient: {avg_dice:.4f} ± {std_dice:.4f}\n")
    f.write(f"  Average clDice: {avg_cldice:.4f} ± {std_cldice:.4f}\n")
    f.write(f"  Average Betti-0 Difference: {avg_betti:.2f} ± {std_betti:.2f}\n")
    f.write(f"  Average IoU: {avg_iou:.4f} ± {std_iou:.4f}\n")
    if all_accuracies:
        f.write(f"\nConnectivity Metrics:\n")
        f.write(f"  Average Connectivity Accuracy: {np.mean(all_accuracies):.4f} ± {np.std(all_accuracies):.4f}\n")
        f.write(f"  Min Accuracy: {np.min(all_accuracies):.4f}\n")
        f.write(f"  Max Accuracy: {np.max(all_accuracies):.4f}\n")

# NEW: Export Excel
export_to_excel(connectivity_results, pred_files, original_results, excel_file)

print(f"\n{'='*60}")
print(f"SUMMARY")
print(f"{'='*60}")
print(f"\nOriginal Metrics:")
print(f"  Average Dice Coefficient: {avg_dice:.4f} ± {std_dice:.4f}")
print(f"  Average clDice: {avg_cldice:.4f} ± {std_cldice:.4f}")
print(f"  Average Betti-0 Difference: {avg_betti:.2f} ± {std_betti:.2f}")
print(f"  Average IoU: {avg_iou:.4f} ± {std_iou:.4f}")
if all_accuracies:
    print(f"\nConnectivity Metrics:")
    print(f"  Average Connectivity Accuracy: {np.mean(all_accuracies):.4f} ± {np.std(all_accuracies):.4f}")
    print(f"  Min Accuracy: {np.min(all_accuracies):.4f}")
    print(f"  Max Accuracy: {np.max(all_accuracies):.4f}")

print(f"\nResults saved to:")
print(f"  - Connectivity only: {connectivity_file}")
print(f"  - Combined metrics: {combined_metrics_file}")
print(f"  - Excel coordinates: {excel_file}")
print(f"  - Masks: {masks_dir}")
print("="*60)